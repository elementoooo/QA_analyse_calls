import json
import re
import os
import time
from google import genai
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

client = genai.Client(api_key="AIzaSyBlyznmQPALtTZaRhlf2SchMmuHM_rD76M")

# ── PALETTE ──────────────────────────────────────────────────────────────────
DARK_NAVY   = "1B2A4A"
MID_BLUE    = "2C4A7C"
ACCENT_BLUE = "3E6DB5"
LIGHT_BLUE  = "D6E4F7"
PALE_BLUE   = "EEF4FB"
WHITE       = "FFFFFF"
LIGHT_GREY  = "F5F7FA"
MID_GREY    = "D0D7E3"
DARK_GREY   = "333333"
GREEN       = "1E7E34"
GREEN_BG    = "D4EDDA"
AMBER       = "856404"
AMBER_BG    = "FFF3CD"
RED         = "721C24"
RED_BG      = "F8D7DA"
HI_FG, HI_BG = "276221", "C6EFCE"
MD_FG, MD_BG = "9C6500", "FFEB9C"
LO_FG, LO_BG = "9C0006", "FFC7CE"

# ── STYLE HELPERS ─────────────────────────────────────────────────────────────
def F(hex_): return PatternFill("solid", fgColor=hex_)
def thin(color=MID_GREY): s = Side(style="thin", color=color); return Border(left=s,right=s,top=s,bottom=s)
def btm_border(): return Border(bottom=Side(style="medium", color=DARK_NAVY))

def score_style(score):
    try:
        s = float(score)
        if s >= 8:  return F(HI_BG), Font(bold=True, color=HI_FG, name="Arial", size=10)
        elif s >= 5: return F(MD_BG), Font(bold=True, color=MD_FG, name="Arial", size=10)
        else:        return F(LO_BG), Font(bold=True, color=LO_FG, name="Arial", size=10)
    except: return F(WHITE), Font(bold=True, name="Arial", size=10)

def verdict_style(v):
    v = str(v).lower()
    if any(x in v for x in ("exemplary","strong")): return F(GREEN_BG), Font(bold=True, color=GREEN, name="Arial", size=9)
    elif "average" in v: return F(AMBER_BG), Font(bold=True, color=AMBER, name="Arial", size=9)
    else: return F(RED_BG), Font(bold=True, color=RED, name="Arial", size=9)

def hdr(ws, row, col, val, bg=DARK_NAVY, fg=WHITE, size=10, bold=True, align="left", span=None):
    if span:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, size=size, color=fg, name="Arial")
    c.fill = F(bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True, indent=1 if align=="left" else 0)
    return c

def cell(ws, row, col, val, bold=False, size=10, fg=DARK_GREY, bg=None, align_h="left", wrap=True, border=None, italic=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, italic=italic, size=size, color=fg, name="Arial")
    c.alignment = Alignment(horizontal=align_h, vertical="top", wrap_text=wrap, indent=1 if align_h=="left" else 0)
    if bg: c.fill = F(bg)
    if border: c.border = border
    return c


# ── JSON CLEANING ─────────────────────────────────────────────────────────────
def clean_json(raw, tag=""):
    if not raw: return None
    raw = re.sub(r"```json|```", "", raw).strip()
    for attempt_fn in [
        lambda t: json.loads(t),
        lambda t: json.loads(re.search(r'\{.*\}', t, re.DOTALL).group()),
        lambda t: json.loads(re.sub(r',\s*$','',t.strip()) + ']'*(t.count('[')-t.count(']')) + '}'*(t.count('{')-t.count('}'))),
    ]:
        try: return attempt_fn(raw)
        except: pass
    print(f"  ❌ [{tag}] JSON parse failed:\n{raw[:500]}")
    return None


# ── ANALYSIS PROMPT (tight output) ───────────────────────────────────────────
def analyze_call(transcript, name=""):
    prompt = f"""You are a senior QA analyst at Suraasa. Evaluate this sales call transcript concisely and return ONLY valid JSON — no markdown, no preamble.

SCORING: 0-3=Critical gap | 4-5=Below avg | 6-7=Average | 8-9=Strong | 10=Exemplary

For each of the 6 pillars provide:
- score (0-10 integer)
- analysis: 2 sentences MAX — what happened and why it matters
- coaching_pointers: exactly 2 specific, actionable items (1 sentence each)
- quote: one short direct quote from transcript (≤12 words) that best exemplifies performance on this pillar

PILLARS:
1. opening — greeting, agenda, rapport, tone
2. engagement — discovery questions, listening, uncovering pain points
3. positioning — linking program to prospect's specific needs
4. value — outcomes, ROI, social proof, urgency
5. commitment — exploring constraints, financial readiness
6. closure — objection handling, next steps, follow-up

Return exactly:
{{
  "opening":     {{"score":0,"analysis":"","coaching_pointers":["",""],"quote":""}},
  "engagement":  {{"score":0,"analysis":"","coaching_pointers":["",""],"quote":""}},
  "positioning": {{"score":0,"analysis":"","coaching_pointers":["",""],"quote":""}},
  "value":       {{"score":0,"analysis":"","coaching_pointers":["",""],"quote":""}},
  "commitment":  {{"score":0,"analysis":"","coaching_pointers":["",""],"quote":""}},
  "closure":     {{"score":0,"analysis":"","coaching_pointers":["",""],"quote":""}},
  "strengths":   ["",""],
  "gaps":        ["",""],
  "overall_score": 0.0,
  "verdict": "Exemplary|Strong|Average|Below Average|Poor",
  "summary": "2 sentence max summary"
}}

TRANSCRIPT:
{transcript}"""

    for model in ["gemini-2.5-flash", "gemini-2.5-flash-lite"]:
        for attempt in range(3):
            try:
                print(f"    [{model}] attempt {attempt+1}...")
                resp = client.models.generate_content(
                    model=model, contents=prompt,
                    config={"temperature": 0.15, "max_output_tokens": 3000}
                )
                parsed = clean_json(resp.text.strip(), name)
                if parsed:
                    required = ["opening","engagement","positioning","value","commitment","closure","overall_score","verdict"]
                    if all(k in parsed for k in required):
                        print(f"  ✅ [{model}] Score:{parsed.get('overall_score')} | {parsed.get('verdict')}")
                        return parsed
                    print(f"  ⚠️  Missing keys, retrying...")
            except Exception as e:
                err = str(e)
                if "503" in err or "UNAVAILABLE" in err:
                    wait = (attempt+1)*20
                    print(f"  ⏳ 503 — waiting {wait}s...")
                    time.sleep(wait)
                elif "404" in err or "NOT_FOUND" in err:
                    print(f"  ❌ Model [{model}] not found"); break
                else:
                    print(f"  ❌ [{name}] error: {e}")
    print(f"  ❌ All models exhausted for [{name}]")
    return None


# ── LOAD TRANSCRIPTS ──────────────────────────────────────────────────────────
folder = os.path.expanduser("~/Desktop/Calls_f")
if not os.path.exists(folder):
    print(f"❌ Folder not found: {folder}"); exit(1)

files = sorted(f for f in os.listdir(folder) if f.endswith((".txt",".rtf")))
if not files:
    print(f"❌ No .txt/.rtf files in {folder}"); exit(1)

print(f"📂 {len(files)} transcript(s) found\n")
transcripts, names = [], []
for f in files:
    with open(os.path.join(folder, f), "r", encoding="utf-8", errors="replace") as fh:
        content = fh.read().strip()
    if not content: print(f"  ⚠️  Skipping empty: {f}"); continue
    transcripts.append(content)
    names.append(f.replace(".txt","").replace(".rtf",""))
    print(f"  ✅ {f} ({len(content):,} chars)")


# ── PROCESS CALLS ─────────────────────────────────────────────────────────────
results = []
for i, (t, n) in enumerate(zip(transcripts, names)):
    print(f"\n{'─'*50}\n[{i+1}/{len(transcripts)}] {n}\n{'─'*50}")
    results.append(analyze_call(t, n))

success = sum(1 for r in results if r)
print(f"\n📊 {success}/{len(results)} parsed successfully")
if not success: print("❌ No data — aborting"); exit(1)


# ── WORKBOOK SETUP ────────────────────────────────────────────────────────────
PILLARS = [
    ("opening",     "Opening"),
    ("engagement",  "Engagement"),
    ("positioning", "Positioning"),
    ("value",       "Value Comm."),
    ("commitment",  "Commitment"),
    ("closure",     "Closure"),
]

wb = Workbook()
wb.remove(wb.active)


# ════════════════════════════════════════════════════════════════════════════
#  SHEET 1 — MASTER HEATMAP  (all counsellors × all pillars)
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("📊 Master Heatmap")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "C4"

N = len(results)
PILLAR_COLS = len(PILLARS)
TOTAL_COLS  = 2 + PILLAR_COLS + 2   # Rank | Counsellor | 6 pillars | Overall | Verdict

# Row 1 — title banner
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
t = ws.cell(row=1, column=1, value="  SURAASA  ·  QA Call Performance Heatmap")
t.font = Font(bold=True, size=13, color=WHITE, name="Arial")
t.fill = F(DARK_NAVY)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 28

# Row 2 — pillar group header
ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=2+PILLAR_COLS)
ph = ws.cell(row=2, column=3, value="PILLAR SCORES  (0–10)")
ph.font = Font(bold=True, size=9, color=WHITE, name="Arial")
ph.fill = F(MID_BLUE)
ph.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 16

# Row 3 — column headers
col_headers = ["#", "Counsellor"] + [p[1] for p in PILLARS] + ["Overall", "Verdict"]
col_widths   = [5,   22]           + [12]*PILLAR_COLS          + [10,       16]
for ci, (h, w) in enumerate(zip(col_headers, col_widths)):
    bg = ACCENT_BLUE if ci >= 2 and ci < 2+PILLAR_COLS else DARK_NAVY
    c  = ws.cell(row=3, column=ci+1, value=h)
    c.font = Font(bold=True, size=9, color=WHITE, name="Arial")
    c.fill = F(bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin()
    ws.column_dimensions[get_column_letter(ci+1)].width = w
ws.row_dimensions[3].height = 22

# Data rows — sorted by overall score desc
sorted_idx = sorted(range(N), key=lambda i: float(results[i].get("overall_score",0)) if results[i] else 0, reverse=True)

for rank, idx in enumerate(sorted_idx):
    r = results[idx]
    row = 4 + rank
    row_bg = PALE_BLUE if rank % 2 == 0 else WHITE

    # Rank
    c = ws.cell(row=row, column=1, value=rank+1)
    c.font = Font(bold=True, size=9, color=DARK_GREY, name="Arial"); c.fill = F(row_bg)
    c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thin()

    # Counsellor name
    c = ws.cell(row=row, column=2, value=names[idx])
    c.font = Font(bold=True, size=9, color=DARK_NAVY, name="Arial"); c.fill = F(row_bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1); c.border = thin()

    if r is None:
        for ci in range(2, TOTAL_COLS):
            ec = ws.cell(row=row, column=ci+1, value="ERR")
            ec.fill = F(LO_BG); ec.border = thin()
            ec.alignment = Alignment(horizontal="center", vertical="center")
    else:
        # Pillar scores
        for pi, (pk, _) in enumerate(PILLARS):
            score = r.get(pk, {}).get("score", "–")
            sc = ws.cell(row=row, column=3+pi, value=score)
            sf, ft = score_style(score)
            sc.fill = sf; sc.font = ft
            sc.alignment = Alignment(horizontal="center", vertical="center"); sc.border = thin()

        # Overall
        overall = r.get("overall_score", "–")
        oc = ws.cell(row=row, column=3+PILLAR_COLS, value=overall)
        sf, ft = score_style(overall)
        oc.fill = sf; oc.font = Font(bold=True, size=10, name="Arial", color=ft.color)
        oc.alignment = Alignment(horizontal="center", vertical="center"); oc.border = thin()

        # Verdict
        verdict = r.get("verdict", "–")
        vc = ws.cell(row=row, column=4+PILLAR_COLS, value=verdict)
        vf, vt = verdict_style(verdict)
        vc.fill = vf; vc.font = vt
        vc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); vc.border = thin()

    ws.row_dimensions[row].height = 18

# Score legend below data
leg_row = 4 + N + 1
ws.merge_cells(start_row=leg_row, start_column=1, end_row=leg_row, end_column=TOTAL_COLS)
legend_items = [
    ("■ 8–10 Strong", HI_FG), ("  ■ 5–7 Average", MD_FG), ("  ■ 0–4 Weak", LO_FG)
]
leg_val = "  Score Legend:   " + "   ".join(f"{l}" for l,_ in legend_items)
lc = ws.cell(row=leg_row, column=1, value=leg_val)
lc.font = Font(size=8, color=DARK_GREY, name="Arial")
lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
lc.fill = F(LIGHT_GREY)
ws.row_dimensions[leg_row].height = 14


# ════════════════════════════════════════════════════════════════════════════
#  SHEET 2 — INDIVIDUAL CALL DETAIL  (compact, all calls stacked)
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("🔍 Call Detail")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 10
ws2.column_dimensions["C"].width = 52
ws2.column_dimensions["D"].width = 45

# Banner
ws2.merge_cells("A1:D1")
t2 = ws2.cell(row=1, column=1, value="  SURAASA  ·  Individual Call Detail — All Counsellors")
t2.font = Font(bold=True, size=12, color=WHITE, name="Arial")
t2.fill = F(DARK_NAVY)
t2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws2.row_dimensions[1].height = 26

drow = 2
for idx in sorted_idx:
    r, name = results[idx], names[idx]

    # ── Call header ──
    ws2.merge_cells(start_row=drow, start_column=1, end_row=drow, end_column=4)
    overall = r.get("overall_score","–") if r else "ERR"
    verdict = r.get("verdict","–") if r else "ERR"
    ch = ws2.cell(row=drow, column=1, value=f"  {name}   ·   Score: {overall}/10   ·   {verdict}")
    vf, vt = verdict_style(verdict)
    ch.fill = vf; ch.font = Font(bold=True, size=11, color=vt.color, name="Arial")
    ch.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ch.border = btm_border()
    ws2.row_dimensions[drow].height = 22
    drow += 1

    if r is None:
        ws2.merge_cells(start_row=drow, start_column=1, end_row=drow, end_column=4)
        ws2.cell(row=drow, column=1, value="  ⚠️ Analysis failed").fill = F(RED_BG)
        drow += 2; continue

    # ── Summary row ──
    ws2.merge_cells(start_row=drow, start_column=1, end_row=drow, end_column=4)
    sc = ws2.cell(row=drow, column=1, value=r.get("summary",""))
    sc.font = Font(italic=True, size=9, color=DARK_GREY, name="Arial"); sc.fill = F(PALE_BLUE)
    sc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    ws2.row_dimensions[drow].height = 32
    drow += 1

    # ── Pillar sub-header ──
    for ci, label in enumerate(["Pillar", "Score", "Analysis (+ Coaching)", "Quote from Call"]):
        c = ws2.cell(row=drow, column=ci+1, value=label)
        c.font = Font(bold=True, size=9, color=WHITE, name="Arial"); c.fill = F(ACCENT_BLUE)
        c.alignment = Alignment(horizontal="center" if ci==1 else "left", vertical="center", indent=0 if ci==1 else 1)
        c.border = thin()
    ws2.row_dimensions[drow].height = 16
    drow += 1

    # ── Per-pillar rows ──
    for pi, (pk, plabel) in enumerate(PILLARS):
        data     = r.get(pk, {})
        score    = data.get("score", "–")
        analysis = data.get("analysis", "")
        pointers = data.get("coaching_pointers", [])
        quote    = data.get("quote", "")
        bg       = PALE_BLUE if pi % 2 == 0 else WHITE

        sf, ft = score_style(score)

        # Combined analysis + coaching in col C
        coaching_text = "\n".join(f"▸ {p}" for p in pointers) if pointers else ""
        combined = analysis + ("\n" + coaching_text if coaching_text else "")

        lc = ws2.cell(row=drow, column=1, value=plabel)
        lc.font = Font(bold=True, size=9, color=DARK_NAVY, name="Arial"); lc.fill = F(bg)
        lc.alignment = Alignment(horizontal="left", vertical="top", indent=1); lc.border = thin()

        sc2 = ws2.cell(row=drow, column=2, value=score)
        sc2.fill = sf; sc2.font = Font(bold=True, size=10, name="Arial", color=ft.color)
        sc2.alignment = Alignment(horizontal="center", vertical="top"); sc2.border = thin()

        ac = ws2.cell(row=drow, column=3, value=combined)
        ac.font = Font(size=9, color=DARK_GREY, name="Arial"); ac.fill = F(bg)
        ac.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1); ac.border = thin()

        qc = ws2.cell(row=drow, column=4, value=f'"{quote}"' if quote else "")
        qc.font = Font(italic=True, size=9, color=MID_BLUE, name="Arial")
        qc.fill = F(LIGHT_BLUE if quote else bg)
        qc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1); qc.border = thin()

        ws2.row_dimensions[drow].height = 55
        drow += 1

    # ── Strengths & Gaps inline ──
    strengths = r.get("strengths", [])
    gaps      = r.get("gaps", [])
    s_text = "✅ " + "  |  ✅ ".join(strengths) if strengths else ""
    g_text = "⚠️ " + "  |  ⚠️ ".join(gaps) if gaps else ""

    sc_c = ws2.cell(row=drow, column=1, value="Strengths")
    sc_c.font = Font(bold=True, size=9, color=GREEN, name="Arial"); sc_c.fill = F(GREEN_BG)
    sc_c.alignment = Alignment(horizontal="left", vertical="top", indent=1); sc_c.border = thin()
    ws2.cell(row=drow, column=2, value="").fill = F(GREEN_BG)
    ws2.merge_cells(start_row=drow, start_column=2, end_row=drow, end_column=4)
    sv = ws2.cell(row=drow, column=2, value=s_text)
    sv.font = Font(size=9, color=GREEN, name="Arial"); sv.fill = F(GREEN_BG)
    sv.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1); sv.border = thin()
    ws2.row_dimensions[drow].height = 22
    drow += 1

    gc_c = ws2.cell(row=drow, column=1, value="Key Gaps")
    gc_c.font = Font(bold=True, size=9, color=RED, name="Arial"); gc_c.fill = F(RED_BG)
    gc_c.alignment = Alignment(horizontal="left", vertical="top", indent=1); gc_c.border = thin()
    ws2.merge_cells(start_row=drow, start_column=2, end_row=drow, end_column=4)
    gv = ws2.cell(row=drow, column=2, value=g_text)
    gv.font = Font(size=9, color=RED, name="Arial"); gv.fill = F(RED_BG)
    gv.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1); gv.border = thin()
    ws2.row_dimensions[drow].height = 22
    drow += 1

    # spacer
    ws2.row_dimensions[drow].height = 8
    drow += 1


# ════════════════════════════════════════════════════════════════════════════
#  SHEET 3 — COACHING TRACKER  (dense, action-oriented)
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("📋 Coaching Tracker")
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = "A3"

ws3.merge_cells("A1:G1")
t3 = ws3.cell(row=1, column=1, value="  SURAASA  ·  Coaching Action Tracker")
t3.font = Font(bold=True, size=12, color=WHITE, name="Arial")
t3.fill = F(DARK_NAVY)
t3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws3.row_dimensions[1].height = 26

ct_hdrs = ["Counsellor", "Pillar", "Score", "Coaching Action", "Evidence Quote", "Priority", "Status"]
ct_wids = [20,           18,       8,       55,                 38,               10,          12]
for ci, (h, w) in enumerate(zip(ct_hdrs, ct_wids)):
    c = ws3.cell(row=2, column=ci+1, value=h)
    c.font = Font(bold=True, size=9, color=WHITE, name="Arial"); c.fill = F(ACCENT_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = thin()
    ws3.column_dimensions[get_column_letter(ci+1)].width = w
ws3.row_dimensions[2].height = 18

trow = 3
for idx in sorted_idx:
    r, name = results[idx], names[idx]
    if r is None: continue
    row_bg = PALE_BLUE if idx % 2 == 0 else WHITE

    for pk, plabel in PILLARS:
        data     = r.get(pk, {})
        score    = data.get("score", "–")
        pointers = data.get("coaching_pointers", [])
        quote    = data.get("quote", "")
        sf, ft   = score_style(score)

        # Priority label based on score
        try:
            s = float(score)
            priority = "🔴 High" if s < 5 else ("🟡 Med" if s < 8 else "🟢 Low")
        except: priority = "—"

        for pidx, pointer in enumerate(pointers):
            c1 = ws3.cell(row=trow, column=1, value=name if pidx == 0 else "")
            c1.font = Font(bold=True, size=9, color=ACCENT_BLUE, name="Arial"); c1.fill = F(row_bg)
            c1.alignment = Alignment(horizontal="left", vertical="top", indent=1); c1.border = thin()

            c2 = ws3.cell(row=trow, column=2, value=plabel if pidx == 0 else "")
            c2.font = Font(bold=(pidx==0), size=9, color=DARK_GREY, name="Arial"); c2.fill = F(row_bg)
            c2.alignment = Alignment(horizontal="left", vertical="top", indent=1); c2.border = thin()

            c3 = ws3.cell(row=trow, column=3, value=score if pidx == 0 else "")
            c3.fill = sf if pidx == 0 else F(row_bg)
            c3.font = Font(bold=True, size=9, name="Arial", color=ft.color if pidx==0 else DARK_GREY)
            c3.alignment = Alignment(horizontal="center", vertical="top"); c3.border = thin()

            c4 = ws3.cell(row=trow, column=4, value=f"{'①②'[pidx]}  {pointer}")
            c4.font = Font(size=9, color=DARK_GREY, name="Arial"); c4.fill = F(row_bg)
            c4.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1); c4.border = thin()

            c5 = ws3.cell(row=trow, column=5, value=f'"{quote}"' if (pidx==0 and quote) else "")
            c5.font = Font(italic=True, size=8, color=MID_BLUE, name="Arial")
            c5.fill = F(LIGHT_BLUE if pidx==0 and quote else row_bg)
            c5.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1); c5.border = thin()

            c6 = ws3.cell(row=trow, column=6, value=priority if pidx==0 else "")
            c6.font = Font(size=9, name="Arial"); c6.fill = F(row_bg)
            c6.alignment = Alignment(horizontal="center", vertical="top"); c6.border = thin()

            c7 = ws3.cell(row=trow, column=7, value="Pending")
            c7.font = Font(size=8, color=AMBER, name="Arial"); c7.fill = F(AMBER_BG)
            c7.alignment = Alignment(horizontal="center", vertical="center"); c7.border = thin()

            ws3.row_dimensions[trow].height = 32
            trow += 1


# ── SAVE ─────────────────────────────────────────────────────────────────────
out = os.path.expanduser("~/Desktop/suraasa_qa_report.xlsx")
wb.save(out)
print(f"\n✅ Report saved → {out}")