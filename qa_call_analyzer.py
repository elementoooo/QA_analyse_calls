import json
import re
import os
import time
from google import genai
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side
)
from openpyxl.utils import get_column_letter

# -------- INIT CLIENT -------- #
client = genai.Client(api_key="urkey")

# -------- PALETTE -------- #
DARK_NAVY   = "1B2A4A"
MID_BLUE    = "2C4A7C"
ACCENT_BLUE = "3E6DB5"
LIGHT_BLUE  = "D6E4F7"
PALE_BLUE   = "EEF4FB"
WHITE       = "FFFFFF"
LIGHT_GREY  = "F5F7FA"
MID_GREY    = "D0D7E3"
DARK_GREY   = "4A4A4A"
GREEN       = "1E7E34"
GREEN_BG    = "D4EDDA"
AMBER       = "856404"
AMBER_BG    = "FFF3CD"
RED         = "721C24"
RED_BG      = "F8D7DA"
SCORE_HIGH  = "C6EFCE"
SCORE_MID   = "FFEB9C"
SCORE_LOW   = "FFC7CE"
SCORE_HIGH_FT = "276221"
SCORE_MID_FT  = "9C6500"
SCORE_LOW_FT  = "9C0006"


# -------- HELPERS -------- #
def thin_border(top=True, bottom=True, left=True, right=True):
    t = Side(style="thin", color=MID_GREY)
    n = None
    return Border(
        top=t if top else n, bottom=t if bottom else n,
        left=t if left else n, right=t if right else n,
    )

def thick_bottom():
    return Border(bottom=Side(style="medium", color=DARK_NAVY))

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def score_fill(score):
    try:
        s = float(score)
        if s >= 8:   return fill(SCORE_HIGH), Font(bold=True, color=SCORE_HIGH_FT, name="Arial", size=11)
        elif s >= 5: return fill(SCORE_MID),  Font(bold=True, color=SCORE_MID_FT,  name="Arial", size=11)
        else:        return fill(SCORE_LOW),  Font(bold=True, color=SCORE_LOW_FT,  name="Arial", size=11)
    except:
        return fill(WHITE), Font(bold=True, name="Arial", size=11)

def verdict_style(verdict):
    v = str(verdict).lower()
    if "exemplary" in v or "strong" in v:
        return fill(GREEN_BG), Font(bold=True, color=GREEN, name="Arial", size=10)
    elif "average" in v:
        return fill(AMBER_BG), Font(bold=True, color=AMBER, name="Arial", size=10)
    else:
        return fill(RED_BG), Font(bold=True, color=RED, name="Arial", size=10)

def set_cell(ws, row, col, value,
             bold=False, italic=False, size=10, color=DARK_GREY,
             bg=None, align_h="left", align_v="top",
             wrap=True, border=None, name="Arial"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, italic=italic, size=size, color=color, name=name)
    c.alignment = Alignment(horizontal=align_h, vertical=align_v,
                             wrap_text=wrap, indent=1 if align_h == "left" else 0)
    if bg:
        c.fill = fill(bg)
    if border:
        c.border = border
    return c


# -------- CLEAN JSON (robust) -------- #
def clean_json(output, call_name=""):
    if not output:
        print(f"  ⚠️  [{call_name}] Empty response from API")
        return None

    # Strip markdown fences
    output = re.sub(r"```json|```", "", output).strip()

    # Attempt 1: direct parse
    try:
        return json.loads(output)
    except:
        pass

    # Attempt 2: extract first {...} block
    try:
        match = re.search(r'\{.*\}', output, re.DOTALL)
        if match:
            return json.loads(match.group())
    except:
        pass

    # Attempt 3: fix truncated JSON by closing open structures
    try:
        text = output
        opens  = text.count('{') - text.count('}')
        aopens = text.count('[') - text.count(']')
        text = re.sub(r',\s*$', '', text.strip())
        text += ']' * aopens + '}' * opens
        return json.loads(text)
    except:
        pass

    print(f"  ❌ [{call_name}] JSON parse failed. Raw output preview:")
    print("  ---")
    print(output[:800])
    print("  ---")
    return None


# -------- ANALYZE CALL -------- #
def analyze_call(transcript, call_name=""):
    prompt = f"""
You are a senior Quality Assurance analyst at Suraasa, evaluating sales counselor calls against Suraasa's core sales methodology. Your analysis must be rigorous, evidence-based, and actionable.

CRITICAL INSTRUCTION: Return ONLY a valid JSON object. No markdown code fences, no preamble, no explanation, no trailing text. Start your response with {{ and end with }}.

Evaluate the transcript across the 6 pillars of Suraasa's sales framework. For each pillar, provide:
- A score from 0–10 (integer)
- A detailed analysis (4–6 sentences)
- Exactly 3 coaching pointers (specific and actionable)
- Exactly 1 reference_line: a short direct quote (max 15 words) from the transcript

SCORING GUIDE:
0–3: Critical gaps
4–5: Below average
6–7: Average
8–9: Strong
10: Exemplary

PILLARS:
1. opening: Warm greeting, agenda-setting, rapport, tone, call ownership
2. engagement: Discovery questions, active listening, uncovering pain points, avoiding assumption-led pitching
3. positioning: Program tailored to prospect's needs, logical flow from problem to solution
4. value: Success stories, outcomes data, urgency, ROI articulation, handling skepticism
5. commitment: Exploring constraints, financial options, readiness gauging
6. closure: Objection handling, closing technique, next steps, follow-up commitment

Return exactly this JSON shape:
{{
  "opening":     {{"score": 0, "analysis": "string", "coaching_pointers": ["s","s","s"], "reference_line": "string"}},
  "engagement":  {{"score": 0, "analysis": "string", "coaching_pointers": ["s","s","s"], "reference_line": "string"}},
  "positioning": {{"score": 0, "analysis": "string", "coaching_pointers": ["s","s","s"], "reference_line": "string"}},
  "value":       {{"score": 0, "analysis": "string", "coaching_pointers": ["s","s","s"], "reference_line": "string"}},
  "commitment":  {{"score": 0, "analysis": "string", "coaching_pointers": ["s","s","s"], "reference_line": "string"}},
  "closure":     {{"score": 0, "analysis": "string", "coaching_pointers": ["s","s","s"], "reference_line": "string"}},
  "top_strengths": ["string","string","string"],
  "critical_improvements": ["string","string","string"],
  "overall_score": 0.0,
  "final_verdict": "Exemplary | Strong | Average | Below Average | Poor",
  "overall_summary": "string"
}}

Transcript:
{transcript}
"""

    models_to_try = ["gemini-2.5-flash", "gemini-2.5-flash-lite"]

    for model_name in models_to_try:
        for attempt in range(3):  # 3 attempts per model
            try:
                print(f"    🔄 [{model_name}] attempt {attempt + 1}...")

                response = client.models.generate_content(
                    model=model_name,
                    contents=prompt,
                    config={
                        "temperature": 0.2,
                        "max_output_tokens": 4500,
                    }
                )

                raw = response.text.strip()
                print(f"    📥 Response length: {len(raw)} chars")

                parsed = clean_json(raw, call_name)
                if parsed:
                    required = ["opening","engagement","positioning","value","commitment","closure","overall_score","final_verdict"]
                    missing  = [k for k in required if k not in parsed]
                    if missing:
                        print(f"  ⚠️  [{call_name}] Missing keys: {missing}. Retrying...")
                        continue
                    print(f"  ✅ Success with [{model_name}]")
                    return parsed
                else:
                    print(f"  ⚠️  Attempt {attempt+1} failed to parse JSON.")

            except Exception as e:
                err = str(e)
                if "503" in err or "UNAVAILABLE" in err:
                    wait = (attempt + 1) * 20  # 20s, 40s, 60s
                    print(f"  ⏳ 503 overload on [{model_name}] — waiting {wait}s before retry...")
                    time.sleep(wait)
                elif "404" in err or "NOT_FOUND" in err:
                    print(f"  ❌ Model [{model_name}] not found, trying next model...")
                    break  # skip remaining attempts, move to next model
                else:
                    print(f"  ❌ [{call_name}] API ERROR (attempt {attempt+1}): {e}")

        print(f"  ⚠️  All attempts failed for [{model_name}], trying fallback...")

    print(f"  ❌ All models exhausted for [{call_name}]")
    return None


# -------- LOAD TRANSCRIPTS -------- #
folder_path = os.path.expanduser("~/Desktop/Calls_f")

if not os.path.exists(folder_path):
    print(f"❌ Folder not found: {folder_path}")
    exit(1)

transcripts = []
CALL_NAMES  = []

txt_files = sorted([f for f in os.listdir(folder_path) if f.endswith(".txt") or f.endswith(".rtf")])
if not txt_files:
    print(f"❌ No .txt/.rtf files found in {folder_path}")
    exit(1)

print(f"📂 Found {len(txt_files)} transcript(s): {txt_files}\n")

for file in txt_files:
    path = os.path.join(folder_path, file)
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        content = f.read().strip()
    if not content:
        print(f"  ⚠️  Skipping empty file: {file}")
        continue
    transcripts.append(content)
    CALL_NAMES.append(file.replace(".txt", "").replace(".rtf", ""))
    print(f"  ✅ Loaded: {file} ({len(content)} chars)")


# -------- PROCESS ALL CALLS -------- #
all_parsed = []
for i, t in enumerate(transcripts):
    print(f"\n{'='*50}")
    print(f"Processing [{i+1}/{len(transcripts)}]: {CALL_NAMES[i]}")
    print(f"{'='*50}")
    parsed = analyze_call(t, CALL_NAMES[i])
    if parsed:
        print(f"  ✅ Overall: {parsed.get('overall_score')} | Verdict: {parsed.get('final_verdict')}")
    else:
        print(f"  ❌ FAILED — this call will show as ERROR in the report")
    all_parsed.append(parsed)


# -------- CHECK WE HAVE DATA -------- #
successful = sum(1 for p in all_parsed if p is not None)
print(f"\n📊 {successful}/{len(all_parsed)} calls parsed successfully")
if successful == 0:
    print("❌ No data to write — aborting report generation.")
    exit(1)


# -------- BUILD WORKBOOK -------- #
wb = Workbook()
wb.remove(wb.active)

PILLARS = [
    ("opening",     "1. Opening & Call Setup"),
    ("engagement",  "2. Engagement, Probing & Need Assessment"),
    ("positioning", "3. Solution Positioning & Program Alignment"),
    ("value",       "4. Value Communication, Credibility & FOMO"),
    ("commitment",  "5. Commitments & Readiness"),
    ("closure",     "6. Call Control, Closure & Next Steps"),
]


# ════════════════════════════════════════════════════
#  SHEET 1 — SCORECARD OVERVIEW
# ════════════════════════════════════════════════════
ws_sc = wb.create_sheet("📊 Scorecard Overview")
ws_sc.sheet_view.showGridLines = False
ws_sc.freeze_panes = "B3"
N = len(all_parsed)

ws_sc.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + N)
tc = ws_sc.cell(row=1, column=1, value="  SURAASA  ·  QA CALL SCORECARD OVERVIEW")
tc.font      = Font(bold=True, size=14, color=WHITE, name="Arial")
tc.fill      = fill(DARK_NAVY)
tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_sc.row_dimensions[1].height = 32

ws_sc.cell(row=2, column=1, value="Pillar").font = Font(bold=True, size=10, color=WHITE, name="Arial")
ws_sc.cell(row=2, column=1).fill      = fill(MID_BLUE)
ws_sc.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws_sc.cell(row=2, column=1).border    = thin_border()
ws_sc.row_dimensions[2].height = 22

for ci, name in enumerate(CALL_NAMES):
    c = ws_sc.cell(row=2, column=2 + ci, value=name)
    c.font      = Font(bold=True, size=10, color=WHITE, name="Arial")
    c.fill      = fill(ACCENT_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = thin_border()

pillar_labels = [p[1] for p in PILLARS] + ["Overall Score", "Final Verdict"]
for ri, label in enumerate(pillar_labels):
    row = 3 + ri
    bg  = PALE_BLUE if ri % 2 == 0 else WHITE
    lc  = ws_sc.cell(row=row, column=1, value=label)
    lc.font      = Font(bold=(label in ["Overall Score", "Final Verdict"]), size=10, color=DARK_GREY, name="Arial")
    lc.fill      = fill(bg)
    lc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    lc.border    = thin_border()
    ws_sc.row_dimensions[row].height = 20

    for ci, parsed in enumerate(all_parsed):
        col = 2 + ci
        if parsed is None:
            c = ws_sc.cell(row=row, column=col, value="ERROR")
            c.border = thin_border()
            continue
        if ri < len(PILLARS):
            key   = PILLARS[ri][0]
            score = parsed.get(key, {}).get("score", "–")
            sc    = ws_sc.cell(row=row, column=col, value=score)
            sf, ft = score_fill(score)
            sc.fill      = sf;  sc.font = ft
            sc.alignment = Alignment(horizontal="center", vertical="center")
            sc.border    = thin_border()
        elif label == "Overall Score":
            val = parsed.get("overall_score", "–")
            sc  = ws_sc.cell(row=row, column=col, value=val)
            sf, ft = score_fill(val)
            sc.fill      = sf
            sc.font      = Font(bold=True, size=11, name="Arial", color=ft.color)
            sc.alignment = Alignment(horizontal="center", vertical="center")
            sc.border    = thin_border()
        else:
            verdict = parsed.get("final_verdict", "–")
            sc      = ws_sc.cell(row=row, column=col, value=verdict)
            vf, vt  = verdict_style(verdict)
            sc.fill      = vf;  sc.font = vt
            sc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sc.border    = thin_border()

ws_sc.column_dimensions["A"].width = 38
for ci in range(N):
    ws_sc.column_dimensions[get_column_letter(2 + ci)].width = 18


# ════════════════════════════════════════════════════
#  ONE DETAIL SHEET PER CALL
# ════════════════════════════════════════════════════
for call_idx, (parsed, call_name) in enumerate(zip(all_parsed, CALL_NAMES)):
    ws = wb.create_sheet(f"🔍 {call_name}")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 80

    row = 1
    ws.merge_cells(f"A{row}:B{row}")
    tc = ws.cell(row=row, column=1, value=f"  SURAASA  ·  QA Analysis  ·  {call_name}")
    tc.font      = Font(bold=True, size=13, color=WHITE, name="Arial")
    tc.fill      = fill(DARK_NAVY)
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 30
    row += 1

    if parsed is None:
        ws.cell(row=row, column=1, value="⚠️  Analysis failed — JSON parse error.")
        continue

    overall = parsed.get("overall_score", "–")
    verdict = parsed.get("final_verdict", "–")
    vf, vt  = verdict_style(verdict)
    ws.merge_cells(f"A{row}:B{row}")
    vc = ws.cell(row=row, column=1, value=f"  Overall Score: {overall} / 10     |     Verdict: {verdict}")
    vc.font      = Font(bold=True, size=11, color=vt.color, name="Arial")
    vc.fill      = vf
    vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24
    row += 1

    ws.merge_cells(f"A{row}:B{row}")
    sh = ws.cell(row=row, column=1, value="  EXECUTIVE SUMMARY")
    sh.font      = Font(bold=True, size=10, color=WHITE, name="Arial")
    sh.fill      = fill(MID_BLUE)
    sh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20
    row += 1

    ws.merge_cells(f"A{row}:B{row}")
    sc = ws.cell(row=row, column=1, value=parsed.get("overall_summary", ""))
    sc.font      = Font(italic=True, size=10, color=DARK_GREY, name="Arial")
    sc.fill      = fill(PALE_BLUE)
    sc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    ws.row_dimensions[row].height = 60
    row += 1

    ws.row_dimensions[row].height = 6
    row += 1

    for pk, plabel in PILLARS:
        data      = parsed.get(pk, {})
        score     = data.get("score", "–")
        analysis  = data.get("analysis", "")
        pointers  = data.get("coaching_pointers", [])
        ref_line  = data.get("reference_line", "")

        ws.merge_cells(f"A{row}:B{row}")
        ph = ws.cell(row=row, column=1, value=f"  {plabel}   ·   Score: {score} / 10")
        ph.font      = Font(bold=True, size=10, color=WHITE, name="Arial")
        ph.fill      = fill(ACCENT_BLUE)
        ph.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ph.border    = thick_bottom()
        ws.row_dimensions[row].height = 22
        row += 1

        sf, ft = score_fill(score)
        ws.cell(row=row, column=1, value="Score").font      = Font(bold=True, size=9, color=DARK_GREY, name="Arial")
        ws.cell(row=row, column=1).fill      = fill(LIGHT_GREY)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=1).border    = thin_border()
        sc2 = ws.cell(row=row, column=2, value=score)
        sc2.fill = sf;  sc2.font = Font(bold=True, size=12, name="Arial", color=ft.color)
        sc2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        sc2.border    = thin_border()
        ws.row_dimensions[row].height = 20
        row += 1

        al = ws.cell(row=row, column=1, value="Analysis")
        al.font = Font(bold=True, size=9, color=WHITE, name="Arial");  al.fill = fill(MID_BLUE)
        al.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        al.border    = thin_border()
        av = ws.cell(row=row, column=2, value=analysis)
        av.font = Font(size=10, color=DARK_GREY, name="Arial");  av.fill = fill(WHITE)
        av.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        av.border    = thin_border()
        ws.row_dimensions[row].height = 80
        row += 1

        if ref_line:
            rl = ws.cell(row=row, column=1, value="📎 From the Call")
            rl.font = Font(bold=True, size=9, color=ACCENT_BLUE, name="Arial");  rl.fill = fill(LIGHT_BLUE)
            rl.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            rl.border    = thin_border()
            rv = ws.cell(row=row, column=2, value=f'"{ref_line}"')
            rv.font = Font(italic=True, size=10, color=MID_BLUE, name="Arial");  rv.fill = fill(LIGHT_BLUE)
            rv.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
            rv.border    = thin_border()
            ws.row_dimensions[row].height = 30
            row += 1

        cl = ws.cell(row=row, column=1, value="Coaching\nPointers")
        cl.font = Font(bold=True, size=9, color=WHITE, name="Arial");  cl.fill = fill(DARK_NAVY)
        cl.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        cl.border    = thin_border()
        pointer_text = "\n".join(f"  {['①','②','③'][idx]}  {p}" for idx, p in enumerate(pointers))
        cv = ws.cell(row=row, column=2, value=pointer_text)
        cv.font = Font(size=10, color=DARK_GREY, name="Arial");  cv.fill = fill(LIGHT_GREY)
        cv.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        cv.border    = thin_border()
        ws.row_dimensions[row].height = max(18 * len(pointers), 55)
        row += 1

        ws.row_dimensions[row].height = 8
        row += 1

    ws.merge_cells(f"A{row}:B{row}")
    si = ws.cell(row=row, column=1, value="  STRENGTHS  &  CRITICAL IMPROVEMENTS")
    si.font      = Font(bold=True, size=10, color=WHITE, name="Arial")
    si.fill      = fill(MID_BLUE)
    si.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22
    row += 1

    strengths    = parsed.get("top_strengths", [])
    improvements = parsed.get("critical_improvements", [])

    sh1 = ws.cell(row=row, column=1, value="✅  Top Strengths")
    sh1.font = Font(bold=True, size=9, color=GREEN, name="Arial");  sh1.fill = fill(GREEN_BG)
    sh1.alignment = Alignment(horizontal="center", vertical="center");  sh1.border = thin_border()
    sh2 = ws.cell(row=row, column=2, value="⚠️  Critical Improvements")
    sh2.font = Font(bold=True, size=9, color=RED, name="Arial");  sh2.fill = fill(RED_BG)
    sh2.alignment = Alignment(horizontal="center", vertical="center");  sh2.border = thin_border()
    ws.row_dimensions[row].height = 20
    row += 1

    for idx in range(max(len(strengths), len(improvements))):
        s_val = f"  {['①','②','③'][idx]}  {strengths[idx]}"    if idx < len(strengths)    else ""
        i_val = f"  {['①','②','③'][idx]}  {improvements[idx]}" if idx < len(improvements) else ""
        sc3 = ws.cell(row=row, column=1, value=s_val)
        sc3.font = Font(size=10, color=GREEN, name="Arial");  sc3.fill = fill(GREEN_BG if idx%2==0 else WHITE)
        sc3.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        sc3.border    = thin_border()
        ic = ws.cell(row=row, column=2, value=i_val)
        ic.font = Font(size=10, color=RED, name="Arial");  ic.fill = fill(RED_BG if idx%2==0 else WHITE)
        ic.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        ic.border    = thin_border()
        ws.row_dimensions[row].height = 45
        row += 1


# ════════════════════════════════════════════════════
#  SHEET 3 — COACHING TRACKER
# ════════════════════════════════════════════════════
ws_ct = wb.create_sheet("📋 Coaching Tracker")
ws_ct.sheet_view.showGridLines = False
ws_ct.freeze_panes = "A3"

ws_ct.merge_cells("A1:F1")
ct_title = ws_ct.cell(row=1, column=1, value="  SURAASA  ·  Coaching Tracker — All Calls")
ct_title.font      = Font(bold=True, size=13, color=WHITE, name="Arial")
ct_title.fill      = fill(DARK_NAVY)
ct_title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_ct.row_dimensions[1].height = 30

ct_headers = ["Call", "Pillar", "Score", "Coaching Pointer", "Reference Line", "Status"]
ct_widths   = [12, 36, 10, 60, 45, 14]
for ci, (h, w) in enumerate(zip(ct_headers, ct_widths)):
    c = ws_ct.cell(row=2, column=ci+1, value=h)
    c.font      = Font(bold=True, size=10, color=WHITE, name="Arial")
    c.fill      = fill(ACCENT_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = thin_border()
    ws_ct.column_dimensions[get_column_letter(ci+1)].width = w
ws_ct.row_dimensions[2].height = 22

ct_row = 3
for call_idx, (parsed, call_name) in enumerate(zip(all_parsed, CALL_NAMES)):
    if parsed is None:
        continue
    row_bg = PALE_BLUE if call_idx % 2 == 0 else WHITE
    for pk, plabel in PILLARS:
        data     = parsed.get(pk, {})
        score    = data.get("score", "–")
        pointers = data.get("coaching_pointers", [])
        ref_line = data.get("reference_line", "")
        sf, ft   = score_fill(score)

        for pidx, pointer in enumerate(pointers):
            c1 = ws_ct.cell(row=ct_row, column=1, value=call_name if pidx==0 else "")
            c1.font = Font(bold=True, size=10, color=ACCENT_BLUE, name="Arial");  c1.fill = fill(row_bg)
            c1.alignment = Alignment(horizontal="center", vertical="top");  c1.border = thin_border()

            c2 = ws_ct.cell(row=ct_row, column=2, value=plabel if pidx==0 else "")
            c2.font = Font(size=10, color=DARK_GREY, name="Arial", bold=(pidx==0));  c2.fill = fill(row_bg)
            c2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
            c2.border = thin_border()

            c3 = ws_ct.cell(row=ct_row, column=3, value=score if pidx==0 else "")
            c3.fill = sf if pidx==0 else fill(row_bg)
            c3.font = Font(bold=True, size=10, name="Arial", color=ft.color if pidx==0 else DARK_GREY)
            c3.alignment = Alignment(horizontal="center", vertical="top");  c3.border = thin_border()

            c4 = ws_ct.cell(row=ct_row, column=4, value=f"{['①','②','③'][pidx]}  {pointer}")
            c4.font = Font(size=10, color=DARK_GREY, name="Arial");  c4.fill = fill(row_bg)
            c4.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
            c4.border = thin_border()

            c5 = ws_ct.cell(row=ct_row, column=5, value=f'"{ref_line}"' if (pidx==0 and ref_line) else "")
            c5.font = Font(italic=True, size=9, color=MID_BLUE, name="Arial")
            c5.fill = fill(LIGHT_BLUE if pidx==0 and ref_line else row_bg)
            c5.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
            c5.border = thin_border()

            c6 = ws_ct.cell(row=ct_row, column=6, value="Pending")
            c6.font = Font(size=9, color=AMBER, name="Arial");  c6.fill = fill(AMBER_BG)
            c6.alignment = Alignment(horizontal="center", vertical="center");  c6.border = thin_border()

            ws_ct.row_dimensions[ct_row].height = 40
            ct_row += 1


# -------- SAVE -------- #
output_path = os.path.expanduser("~/Desktop/gem_qa_report.xlsx")
wb.save(output_path)
print(f"\n✅ Report saved → {output_path}")
